"""
Testes de integração para o sistema de renderização de relatórios.
"""

import json
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from docsync.exceptions import RenderError
from docsync.utils.filter_registry import FilterRegistry
from docsync.utils.renderer import ReportRenderer

# Fixtures


@pytest.fixture
def templates_dir(tmp_path):
    """Cria diretório temporário com templates de teste."""
    templates = tmp_path / "templates"
    templates.mkdir()
    return templates


@pytest.fixture
def output_dir(tmp_path):
    """Cria diretório temporário para outputs."""
    output = tmp_path / "output"
    output.mkdir()
    return output


@pytest.fixture
def sample_data():
    """Dados de exemplo para testes."""
    return {
        "report": {
            "title": "Relatório ESG Q1 2024",
            "period": "Q1 2024",
            "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "version": "1.0.0",
        },
        "metrics": [
            {
                "name": "Emissão de CO2",
                "value": 125.5,
                "target": "100 ton",
                "status": "at_risk",
                "trend": "increasing",
                "unit": "ton",
            },
            {
                "name": "Consumo de Energia",
                "value": 450.75,
                "target": "400 kWh",
                "status": "in_progress",
                "trend": "stable",
                "unit": "kWh",
            },
        ],
        "objectives": [
            {
                "title": "Redução de Emissões",
                "progress": 65,
                "status": "in_progress",
                "responsible": "John Doe",
                "description": "Reduzir emissões em 20%",
            },
        ],
    }


@pytest.fixture
def renderer(templates_dir):
    """Cria instância do renderer com filtros."""
    filters = FilterRegistry()
    return ReportRenderer(templates_dir, filters)


# Testes de Integração


def test_json_rendering(templates_dir, output_dir, renderer, sample_data):
    """Testa renderização de relatório JSON."""
    # Cria template
    template = templates_dir / "test.json.jinja"
    template.write_text(
        """
    {
      "report": {
        "title": "{{ report.title }}",
        "period": "{{ report.period }}",
        "metrics": [
          {% for metric in metrics %}
          {
            "name": "{{ metric.name }}",
            "value": {{ metric.value }},
            "status": "{{ metric.status }}",
            "trend": "{{ metric.trend | format_trend }}"
          }{% if not loop.last %},{% endif %}
          {% endfor %}
        ]
      }
    }
    """,
    )

    # Renderiza
    output_file = output_dir / "test.json"
    renderer.render("test", output_file, sample_data, "json")

    # Valida output
    assert output_file.exists()
    with open(output_file) as f:
        data = json.load(f)

    assert data["report"]["title"] == sample_data["report"]["title"]
    assert len(data["report"]["metrics"]) == len(sample_data["metrics"])


def test_excel_rendering(templates_dir, output_dir, renderer, sample_data):
    """Testa renderização de relatório Excel."""
    # Cria template
    template = templates_dir / "test.xlsx.jinja"
    template.write_text(
        """
    [Overview]
    Title|Period|Version
    {{ report.title }}|{{ report.period }}|{{ report.version }}

    [Metrics]
    Name|Value|Target|Status|Trend
    {% for metric in metrics %}
    {{ metric.name }}|{{ metric.value }}|{{ metric.target }}|{{ metric.status }}|{{ metric.trend }}
    {% endfor %}

    [Objectives]
    Title|Progress|Status|Responsible
    {% for obj in objectives %}
    {{ obj.title }}|{{ obj.progress }}|{{ obj.status }}|{{ obj.responsible }}
    {% endfor %}
    """,
    )

    # Renderiza
    output_file = output_dir / "test.xlsx"
    renderer.render("test", output_file, sample_data, "xlsx")

    # Valida output
    assert output_file.exists()
    wb = load_workbook(output_file)

    assert "Overview" in wb.sheetnames
    assert "Metrics" in wb.sheetnames
    assert "Objectives" in wb.sheetnames

    metrics = wb["Metrics"]
    assert metrics.max_row == len(sample_data["metrics"]) + 2  # header + data


def test_latex_rendering(templates_dir, output_dir, renderer, sample_data):
    """Testa renderização de relatório LaTeX."""
    # Cria template
    template = templates_dir / "test.tex.jinja"
    template.write_text(
        r"""
    \documentclass{article}
    \begin{document}

    \title{ {{- report.title -}} }
    \date{ {{- report.generated_at -}} }
    \maketitle

    \section{Métricas}
    {% for metric in metrics %}
    \subsection{ {{- metric.name -}} }
    \begin{itemize}
        \item Valor: {{ metric.value }}
        \item Meta: {{ metric.target }}
        \item Status: {{ metric.status | format_status }}
    \end{itemize}
    {% endfor %}

    \end{document}
    """,
    )

    # Renderiza
    output_file = output_dir / "test.tex"
    renderer.render("test", output_file, sample_data, "tex")

    # Valida output
    assert output_file.exists()
    content = output_file.read_text()

    assert r"\title{" in content
    assert r"\section{Métricas}" in content
    for metric in sample_data["metrics"]:
        assert metric["name"] in content


def test_invalid_template(templates_dir, output_dir, renderer, sample_data):
    """Testa erro com template inválido."""
    with pytest.raises(RenderError):
        renderer.render("nonexistent", output_dir / "test.json", sample_data)


def test_invalid_json_template(templates_dir, output_dir, renderer, sample_data):
    """Testa erro com template JSON mal formatado."""
    # Cria template inválido
    template = templates_dir / "invalid.json.jinja"
    template.write_text(
        """
    {
      "report": {
        "metrics": [
          {% for metric in metrics %}
          { "value": {{ metric.value }} }
          {% if not loop.last %},{% endif %}  {# Erro: vírgula após último item #}
          {% endfor %}
        ],
      }  {# Erro: vírgula extra #}
    }
    """,
    )

    with pytest.raises(RenderError):
        renderer.render("invalid", output_dir / "test.json", sample_data, "json")


def test_excel_data_validation(templates_dir, output_dir, renderer, sample_data):
    """Testa validação de dados no Excel."""
    # Modifica dados com valores problemáticos
    data = sample_data.copy()
    data["metrics"][0]["value"] = float("inf")  # Valor não suportado

    # Cria template
    template = templates_dir / "test.xlsx.jinja"
    template.write_text(
        """
    [Metrics]
    Name|Value
    {% for metric in metrics %}
    {{ metric.name }}|{{ metric.value }}
    {% endfor %}
    """,
    )

    # Renderiza deve falhar graciosamente
    output_file = output_dir / "test.xlsx"
    with pytest.raises(RenderError):
        renderer.render("test", output_file, data, "xlsx")


def test_format_handlers(templates_dir, output_dir, renderer, sample_data):
    """Testa registro de handlers de formato customizados."""
    # Tenta formato não suportado
    with pytest.raises(RenderError):
        renderer.render("test", output_dir / "test.doc", sample_data, "doc")

    # Registra novo handler
    def render_txt(template, output_path, data):
        content = template.render(**data)
        Path(output_path).write_text(content)

    renderer._format_handlers["txt"] = render_txt

    # Cria template
    template = templates_dir / "test.txt.jinja"
    template.write_text(
        """
    Relatório: {{ report.title }}
    Período: {{ report.period }}

    Métricas:
    {% for metric in metrics %}
    - {{ metric.name }}: {{ metric.value }} {{ metric.unit }}
    {% endfor %}
    """,
    )

    # Testa novo formato
    output_file = output_dir / "test.txt"
    renderer.render("test", output_file, sample_data, "txt")

    assert output_file.exists()
    content = output_file.read_text()
    assert sample_data["report"]["title"] in content
